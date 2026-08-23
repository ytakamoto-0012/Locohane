"""read_excel.pyの`--query-json`に`list_charts`opを追加した回帰テスト。

背景（2026-08-23）: excel-vbaマクロブック作成タスクのverifierが、グラフの
存在確認をexcel-renderの画像確認だけに頼っており、画像にグラフが写って
いるか判別できず長時間の推論ループ（ThinkingLoopDetected）に陥った。その後
workerも`read_excel.py --query-json '[{"op": "list_charts"}]'`を試したが
`未対応のqueryです`で失敗した（対応していたのは`group_by`/`list_images`の
み）。グラフの存在・タイトル・種類・位置を画像化せず確認できる`list_charts`
queryを追加した。
"""

import sys
from pathlib import Path

import openpyxl
from openpyxl.chart import BarChart, LineChart, Reference

_SCRIPTS_DIR = Path(__file__).resolve().parent.parent / "skills" / "excel-read" / "scripts"
if str(_SCRIPTS_DIR) not in sys.path:
    sys.path.insert(0, str(_SCRIPTS_DIR))

from read_excel import _read_xlsx  # noqa: E402


def test_list_charts_reports_type_title_and_anchor(tmp_path):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    for i, v in enumerate([1, 2, 3], start=1):
        ws.cell(row=i, column=1, value=v)

    chart = LineChart()
    chart.title = "月別収支推移"
    chart.add_data(Reference(ws, min_col=1, min_row=1, max_row=3))
    ws.add_chart(chart, "F1")

    path = tmp_path / "book.xlsx"
    wb.save(path)

    result = _read_xlsx(path, sheet_arg="Sheet1", offset=0, limit=200, data_only=False, queries=[{"op": "list_charts"}])

    items = result["query_results"][0]["items"]
    assert len(items) == 1
    assert items[0]["chart_index"] == 0
    assert items[0]["type"] == "line"
    assert items[0]["title"] == "月別収支推移"
    assert items[0]["anchor"] == "F1"


def test_list_charts_empty_sheet_returns_empty_list(tmp_path):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    ws["A1"] = "x"
    path = tmp_path / "book.xlsx"
    wb.save(path)

    result = _read_xlsx(path, sheet_arg="Sheet1", offset=0, limit=200, data_only=False, queries=[{"op": "list_charts"}])

    assert result["query_results"][0]["items"] == []


def test_list_charts_untitled_chart_has_null_title(tmp_path):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    for i, v in enumerate([1, 2, 3], start=1):
        ws.cell(row=i, column=1, value=v)

    chart = BarChart()
    chart.add_data(Reference(ws, min_col=1, min_row=1, max_row=3))
    ws.add_chart(chart, "D1")

    path = tmp_path / "book.xlsx"
    wb.save(path)

    result = _read_xlsx(path, sheet_arg="Sheet1", offset=0, limit=200, data_only=False, queries=[{"op": "list_charts"}])

    items = result["query_results"][0]["items"]
    assert items[0]["type"] == "bar"
    assert items[0]["title"] is None
