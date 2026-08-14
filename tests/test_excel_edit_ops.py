"""excel-editスキル（skills/excel-edit/scripts/_ops.py）の書式消失バグ回帰テスト。

背景: annual_schedule.xlsxの週間予定表シートで、月・週（A/B列）だけを
rowsに渡してdelete_rows後にset_rangeを呼んだところ、渡さなかったC〜G列の
既存データ・書式（罫線・背景色）が無警告で失われた。

1. set_cellはstyle未指定なら新規セルへ何も書式を付けない（現行仕様の固定化）。
2. set_cellにinherit_style:trueを指定すると隣接セルの書式を引き継ぐ
   （extract_style/apply_styleの既存の組み合わせを使う）。
3. set_range/insert_row_groupが、シート既存の最大列数より少ない列数の
   rowsで書き込む場合、標準エラーに警告を出す。
"""

import sys
from pathlib import Path

import openpyxl
import pytest

_SCRIPTS_DIR = Path(__file__).resolve().parent.parent / "skills" / "excel-edit" / "scripts"
if str(_SCRIPTS_DIR) not in sys.path:
    sys.path.insert(0, str(_SCRIPTS_DIR))

from _ops import apply_op  # noqa: E402


def _make_wb_with_styled_row():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    thin_style = {"border": "thin", "fill_color": "F2F2F2"}
    apply_op(wb, {"op": "set_cell", "sheet": "Sheet1", "cell": "A1", "value": "月", "style": thin_style})
    apply_op(wb, {"op": "set_cell", "sheet": "Sheet1", "cell": "B1", "value": "週", "style": thin_style})
    return wb, ws


def test_set_cell_without_style_leaves_new_cell_unstyled():
    wb, ws = _make_wb_with_styled_row()
    apply_op(wb, {"op": "set_cell", "sheet": "Sheet1", "cell": "C1", "value": "備考"})

    cell = ws["C1"]
    assert cell.value == "備考"
    assert cell.border.left.style is None
    assert cell.fill.fgColor.rgb in (None, "00000000")


def test_set_cell_inherit_style_copies_neighbor_style():
    wb, ws = _make_wb_with_styled_row()
    apply_op(
        wb,
        {
            "op": "set_cell",
            "sheet": "Sheet1",
            "cell": "C1",
            "value": "備考",
            "inherit_style": True,
            "inherit_style_from": "left",
        },
    )

    cell = ws["C1"]
    assert cell.value == "備考"
    assert cell.border.left.style == "thin"
    assert cell.fill.fgColor.rgb == "00F2F2F2"


def test_set_range_fewer_columns_than_existing_sheet_warns(capsys: pytest.CaptureFixture[str]):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    # 既存7列のヘッダー行を作る（月・週・準備・打合せ・当日・片付け・備考）。
    apply_op(
        wb,
        {
            "op": "set_range",
            "sheet": "Sheet1",
            "start_cell": "A1",
            "rows": [["月", "週", "準備", "打合せ", "当日", "片付け", "備考"]],
        },
    )
    capsys.readouterr()  # 上の呼び出し分の出力は今回の検証対象外なので読み捨てる

    # A/B列（2列）だけをrowsに渡して書き戻す（今回の実害と同じ操作パターン）。
    apply_op(
        wb,
        {
            "op": "set_range",
            "sheet": "Sheet1",
            "start_cell": "A2",
            "rows": [["4月", "第1週4月"]],
        },
    )

    captured = capsys.readouterr()
    assert "警告" in captured.err
    assert "7列目" in captured.err


def test_set_range_matching_columns_does_not_warn(capsys: pytest.CaptureFixture[str]):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    apply_op(
        wb,
        {
            "op": "set_range",
            "sheet": "Sheet1",
            "start_cell": "A1",
            "rows": [["月", "週"], ["4月", "第1週4月"]],
        },
    )

    captured = capsys.readouterr()
    assert captured.err == ""
