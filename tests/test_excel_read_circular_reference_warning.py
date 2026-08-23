"""read_excel.pyの循環参照検知（warnings）の回帰テスト。

背景（2026-08-23）: excel-vbaマクロブック作成タスクで、月別収支表シートの
合計セルN8に`=SUM(B8:N8)`（N8自身を含む範囲）という数式が書き込まれ、
Excelで循環参照エラーになっていた。生成直後のverifierサブエージェントは
「数式が正しく設定されているか」を検証項目に含んでいたにもかかわらず、
これを見逃した（`agents/verifier.md`のチェックリストに循環参照の項目が
無く、`read_excel.py`側にも機械的な検知が無かったため）。

列幅超過警告と同じ`warnings`機構に、数式が自身のセルを参照範囲に含む
場合の警告を追加した。これにより、verifierが`warnings`を確認する既存の
手順（`agents/verifier.md`）に自動的に乗る。
"""

import sys
from pathlib import Path

import openpyxl
import pytest

_SCRIPTS_DIR = Path(__file__).resolve().parent.parent / "skills" / "excel-read" / "scripts"
if str(_SCRIPTS_DIR) not in sys.path:
    sys.path.insert(0, str(_SCRIPTS_DIR))

from read_excel import _formula_self_references, _read_xlsx  # noqa: E402


class TestFormulaSelfReferences:
    def test_sum_range_including_own_cell_is_detected(self):
        assert _formula_self_references("=SUM(B8:N8)", own_col=14, own_row=8) is True

    def test_sum_range_excluding_own_cell_is_not_detected(self):
        assert _formula_self_references("=SUM(B8:M8)", own_col=14, own_row=8) is False

    def test_single_cell_self_reference_is_detected(self):
        assert _formula_self_references("=A1+1", own_col=1, own_row=1) is True

    def test_chained_non_circular_reference_is_not_detected(self):
        # C25 = B25+B24 は自身（C25）を範囲に含まない
        assert _formula_self_references("=B25+B24", own_col=3, own_row=25) is False

    def test_sheet_qualified_reference_is_excluded_to_avoid_false_positive(self):
        # シート修飾された参照は判定対象外（別シートの同一座標との衝突を誤検知しない）
        assert _formula_self_references("=SUM('Sheet2'!A1:B8)", own_col=1, own_row=1) is False


class TestReadXlsxCircularReferenceWarning:
    def test_self_referencing_sum_produces_warning(self, tmp_path):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Sheet1"
        ws["B8"] = 1
        ws["N8"] = "=SUM(B8:N8)"
        path = tmp_path / "book.xlsx"
        wb.save(path)

        result = _read_xlsx(path, sheet_arg="Sheet1", offset=0, limit=200, data_only=False)

        assert "warnings" in result
        assert any("循環参照" in w and "N8" in w for w in result["warnings"])

    def test_correct_sum_range_produces_no_circular_warning(self, tmp_path):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Sheet1"
        ws["B8"] = 1
        ws["M8"] = 2
        ws["N8"] = "=SUM(B8:M8)"
        path = tmp_path / "book.xlsx"
        wb.save(path)

        result = _read_xlsx(path, sheet_arg="Sheet1", offset=0, limit=200, data_only=False)

        assert "warnings" not in result or not any("循環参照" in w for w in result.get("warnings", []))

    def test_data_only_mode_does_not_false_positive(self, tmp_path):
        """--data-only指定時は数式でなくキャッシュ値が入るため、循環参照検知は動かない
        （そもそも数式文字列が見えないので誤検知のしようがない、を確認する）。"""
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Sheet1"
        ws["B8"] = 1
        ws["N8"] = "=SUM(B8:N8)"
        path = tmp_path / "book.xlsx"
        wb.save(path)

        result = _read_xlsx(path, sheet_arg="Sheet1", offset=0, limit=200, data_only=True)

        assert "warnings" not in result or not any("循環参照" in w for w in result.get("warnings", []))
