"""既存xlsxの「週間予定表」修正シナリオ用フィクスチャを生成するスクリプト。

evals/cases/system_prompt/ の annual_schedule_week_fix_ambiguous_calendar
ケースが参照する annual_schedule.xlsx を生成する。

本番ログ（app_20260814_001030.log）で実際に観測された不具合を再現する:
「週間予定表」シートのA列（月、結合セル）とB列（週、「第X週Y月」）の組合せが、
実カレンダー通りの週境界と一致しない。かつ、月ごとの行数（結合セルの範囲）は
固定（4,4,4,6,6,8,6,6,7,7,7,9行、fiscal year 4月始まり）だが、これは実際の
月ごとの週数と一致しないため、「行数を変えずに正しい週を割り当てる」ことも
「実カレンダー通りに割り当てる」こともできない、という前提矛盾を含む
（=モデルが自問自答ループに陥った本番事例の直接の原因）。

B列の値は「4行区切りで月が変わる」という単純な（誤った）規則で機械的に
生成しており、実際の行数境界（block_sizes）とはズレるため、本番同様の
ズレパターン（月の最初の数行に前月の週が混入する、重複する週番号が生じる等）
が自然に生じる。

使い方:
    python evals/fixtures/generate_annual_schedule_week_fix_fixture.py
"""

from __future__ import annotations

from pathlib import Path

from openpyxl import Workbook

FIXTURE_ROOT = Path(__file__).resolve().parent / "annual_schedule_week_fix"

# fiscal year（4月始まり）の月順と、各月の固定行数（実データの本番ログに準拠）。
_MONTHS_FISCAL = [4, 5, 6, 7, 8, 9, 10, 11, 12, 1, 2, 3]
_BLOCK_SIZES = [4, 4, 4, 6, 6, 8, 6, 6, 7, 7, 7, 9]


def _build_week_sheet(wb: Workbook) -> None:
    ws = wb.active
    ws.title = "週間予定表"
    ws.append(["月", "週", "準備", "打ち合わせ", "当日", "片付け"])

    row = 2
    for month, size in zip(_MONTHS_FISCAL, _BLOCK_SIZES):
        start_row = row
        for _ in range(size):
            row_index = row - 2  # 0-based data row index
            naive_month = _MONTHS_FISCAL[(row_index // 4) % 12]
            naive_week = (row_index % 4) + 1
            ws.cell(row=row, column=1, value=f"{month}月")
            ws.cell(row=row, column=2, value=f"第{naive_week}週{naive_month}月")
            # C〜F列にも行ごとの固有データを入れる（空欄だと、行の挿入・削除で
            # 既存データの再配置が必要になるという本番同様のジレンマが
            # 再現されず、評価にならないため）。
            ws.cell(row=row, column=3, value=f"資材準備{row}")
            ws.cell(row=row, column=4, value=f"打ち合わせ{row}")
            ws.cell(row=row, column=5, value=f"当日対応{row}")
            ws.cell(row=row, column=6, value=f"片付け{row}")
            row += 1
        end_row = row - 1
        if end_row > start_row:
            ws.merge_cells(f"A{start_row}:A{end_row}")


def _build_month_sheet(wb: Workbook) -> None:
    ws = wb.create_sheet("月間予定表")
    ws.append(["月", "主な行事", "備考"])
    for month in _MONTHS_FISCAL:
        ws.append([f"{month}月", "", ""])


def build_fixture(out_dir: Path = FIXTURE_ROOT) -> Path:
    out_dir.mkdir(parents=True, exist_ok=True)
    wb = Workbook()
    _build_week_sheet(wb)
    _build_month_sheet(wb)
    out_path = out_dir / "annual_schedule.xlsx"
    wb.save(out_path)
    return out_path


def main() -> None:
    path = build_fixture()
    print(f"生成完了: {path}")


if __name__ == "__main__":
    main()
