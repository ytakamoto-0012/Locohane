"""excel系スキル（excel-edit/excel-read/excel-recalc/excel-render/excel-vba-edit/
excel-vba-read）が共有するヘルパー関数。

実体はこのファイルのみ（`office_shared/`配下、SKILL.mdを持たない非スキル
ディレクトリ）。各スキルのscripts/配下のスクリプトはsys.path経由でこの
ファイルを直接importする（1-B 相互import方式、詳細はOFFICE_SKILLS_README.md
参照）。

run_script からは直接実行されない。
"""

from __future__ import annotations

import hashlib
import json
import os
import sys
import shutil
import time as time_module
import unicodedata
from datetime import date, datetime, time
from pathlib import Path


def setup_utf8_stdio() -> None:
    """標準出力/標準エラーをUTF-8に固定する。

    Windows環境ではパイプ経由のPython子プロセスの既定エンコーディングが
    システムのANSIコードページ（例: cp932）になり、run_script側の
    encoding="utf-8" デコードと食い違って日本語が文字化けするため、
    各スクリプトの先頭で必ず呼ぶこと。
    """
    sys.stdout.reconfigure(encoding="utf-8")
    sys.stderr.reconfigure(encoding="utf-8")


def backup_before_overwrite(path: Path) -> Path | None:
    """path が既に存在する場合、上書き直前に同じフォルダへタイムスタンプ付きで
    コピーしてバックアップを作成する。存在しなければ何もせず None を返す
    （新規作成時はバックアップ不要）。
    """
    if not path.exists():
        return None
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    backup_path = path.with_name(f"{path.stem}.bak_{timestamp}{path.suffix}")
    suffix_n = 2
    while backup_path.exists():
        backup_path = path.with_name(f"{path.stem}.bak_{timestamp}_{suffix_n}{path.suffix}")
        suffix_n += 1
    shutil.copy2(path, backup_path)
    return backup_path


def _excel_locks_registry():
    """自セッション専用サンドボックス（`_tmp_<thread_id>/excel_locks/pids.json`）内の
    Excel COMプロセスPIDレジストリのパスと、それを操作するための`path_memory`
    モジュールを返す。

    以前はレジストリを対象ファイルと同じフォルダ（ユーザーの作業ディレクトリ、
    LLMが自由に読み書きできる場所）に置いていたが、以下2点の問題があった
    （2026-08-23 issue対応）:
    1. `release_excel_pid`がプロセスの実終了を確認せず無条件に呼ばれていたため、
       `Close()`/`Quit()`が実効を持たなくてもPID記録だけが消え、実在する
       オーファンプロセスを追跡できなくなっていた。
    2. 記録場所がLLMの自由な読み書き対象だったため、「追跡対象PIDのみを
       終了する」というガイダンスがプローズ（自然文）の注意書きに留まり、
       低パラメータモデルには確実に守らせられなかった（実際、SKILL.mdに
       同旨の注意があったにもかかわらず無差別`taskkill`が実行された）。

    `path_memory.exec_tmp_dir()`が使う既存の`_tmp_<thread_id>`サンドボックス
    （他セッションから読み取り不可・保持日数ベースで自動削除）を流用することで、
    「自セッションが記録した対象だけが操作対象になる」ことをコードレベルで
    保証する。

    Returns:
        `(registry_path, path_memory_module)`のタプル。`AGENT_SRC_DIR`未設定・
        `path_memory`のimport失敗時（run_script以外から直接実行された場合等）は
        `None`（フェイルオープン。既存の`record_excel_pid`と同じ方針）。
    """
    src_dir = os.environ.get("AGENT_SRC_DIR")
    if not src_dir:
        return None
    if src_dir not in sys.path:
        sys.path.insert(0, src_dir)
    try:
        import path_memory
    except ImportError:
        return None
    registry_path = path_memory.exec_tmp_dir("excel_locks") / "pids.json"
    return registry_path, path_memory


def _load_json_list(path: Path) -> list[dict]:
    if not path.is_file():
        return []
    try:
        data = json.loads(path.read_text(encoding="utf-8"))
    except (json.JSONDecodeError, OSError):
        return []
    return data if isinstance(data, list) else []


def _save_json_list(path: Path, entries: list[dict]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(json.dumps(entries, ensure_ascii=False, indent=2), encoding="utf-8")


def is_process_running(pid: int) -> bool:
    """指定PIDのプロセスがまだ実行中かどうかを確認する（Windows専用）。

    `workbook.Close()`/`excel.Quit()`がCOM側で例外を出さないことと、実際の
    OSプロセスが終了していることは別物（2026-08-23
    issue: この確認をせず`release_excel_pid`を無条件に呼んでいたため、
    生き残っていたオーファンプロセスの記録だけが消えていた）。

    判定自体に失敗した場合（権限不足・対象PIDが既に別プロセスに再利用
    されている等）は、`record_excel_pid`のフェイルオープン方針とは逆に
    **安全側＝「生存している」扱い**でTrueを返す。誤って記録を消して
    追跡不能にしてしまうより、記録が残り続ける（次回`terminate_tracked_processes`
    実行時に再判定される）副作用の方が軽微なため。
    """
    try:
        import win32api
        import win32con
        import win32process

        handle = win32api.OpenProcess(win32con.PROCESS_QUERY_LIMITED_INFORMATION, False, pid)
    except Exception:
        return True
    try:
        exit_code = win32process.GetExitCodeProcess(handle)
        return exit_code == win32con.STILL_ACTIVE
    except Exception:
        return True
    finally:
        try:
            win32api.CloseHandle(handle)
        except Exception:
            pass


def wait_for_process_exit(pid: int, timeout_seconds: float = 3.0, poll_interval: float = 0.2) -> bool:
    """指定PIDが`timeout_seconds`以内に終了するのを待つ。終了を確認できればTrue。

    `excel.Quit()`はCOM側への終了要求の送出であり、実プロセスの終了までは
    一瞬のラグがありうるため、即座に`is_process_running`を1回呼ぶのではなく
    短時間ポーリングしてから最終判定する。
    """
    deadline = time_module.monotonic() + timeout_seconds
    while True:
        if not is_process_running(pid):
            return True
        if time_module.monotonic() >= deadline:
            return False
        time_module.sleep(poll_interval)


def record_excel_pid(target_path, excel) -> int | None:
    """COM経由で起動したExcel.ApplicationのPIDを取得し、自セッションの
    PIDレジストリ（`_tmp_<thread_id>/excel_locks/pids.json`）へ追記する。

    run_macroがハングしてrun_scriptの外部タイムアウト（既定300秒）で
    Pythonプロセスごと強制終了されると、finally節が実行されず
    EXCEL.EXEだけが対象ファイルを開いたまま残留する。その際に対象PIDを
    確実に特定して`terminate_tracked_processes`で後始末できるよう、
    Excel起動直後に記録しておく（正常終了時はプロセスの実終了を確認した
    上でrelease_excel_pidが取り除く）。
    Application.HwndはVisible=Falseでも有効なウィンドウハンドルを返すため、
    そこからGetWindowThreadProcessIdでPIDを引ける。取得に失敗した場合
    （pywin32未導入、Hwnd取得不可等）はNoneを返し、呼び出し側は記録を諦める
    （フェイルオープン。記録の失敗で本来の編集処理までブロックしないため）。
    レジストリへの記録自体ができない場合（AGENT_SRC_DIR未設定等）でも、
    取得できたPID自体は返す（呼び出し側のwait_for_process_exit等には使える）。
    """
    try:
        import win32process

        hwnd = excel.Hwnd
        if not hwnd:
            return None
        _, pid = win32process.GetWindowThreadProcessId(hwnd)
    except Exception:
        return None
    resolved = _excel_locks_registry()
    if resolved is None:
        return pid
    registry_path, pm = resolved
    lock_path = registry_path.parent / f"{registry_path.name}.lock"
    with pm._locked(lock_path) as acquired:
        if not acquired:
            return pid
        entries = _load_json_list(registry_path)
        if not any(e.get("pid") == pid for e in entries):
            entries.append(
                {
                    "pid": pid,
                    "target_path": str(Path(target_path).resolve()),
                    "recorded_at": datetime.now().isoformat(timespec="seconds"),
                }
            )
        _save_json_list(registry_path, entries)
    return pid


def release_excel_pid(pid: int | None) -> None:
    """record_excel_pidで記録したPIDをレジストリから取り除く。

    呼び出し側（edit_vba.py等のfinally節）は、`wait_for_process_exit`等で
    プロセスの実終了を確認できた場合のみこれを呼ぶこと（2026-08-23
    issue対応。以前は無条件で呼んでいたため、`Quit()`が実効を持たなかった
    ケースでもPID記録だけが消え、以後追跡不能になっていた）。
    記録が残っていない（pidがNone、レジストリ自体が無い等）場合は何もしない。
    """
    if pid is None:
        return
    resolved = _excel_locks_registry()
    if resolved is None:
        return
    registry_path, pm = resolved
    lock_path = registry_path.parent / f"{registry_path.name}.lock"
    with pm._locked(lock_path) as acquired:
        if not acquired:
            return
        entries = _load_json_list(registry_path)
        remaining = [e for e in entries if e.get("pid") != pid]
        if remaining:
            _save_json_list(registry_path, remaining)
        elif registry_path.exists():
            registry_path.unlink()


def terminate_tracked_processes() -> list[dict]:
    """自セッションが起動したExcel COMプロセスのうち、まだ生存しているものを
    強制終了する。

    `run_macro`のハング等でrun_scriptの外部タイムアウトによりPythonプロセス
    ごと強制終了されると、`edit_vba.py`等のfinally節が実行されずEXCEL.EXEだけが
    対象ファイルを開いたまま残留することがある（issue/20260822_212800）。
    この関数は、自セッションの`_tmp_<thread_id>/excel_locks/pids.json`に
    記録されたPIDのみを対象に、生存確認の上で直接`TerminateProcess`する
    （`taskkill`のようなシェルコマンドやイメージ名フィルタは一切使わない。
    LLMがPID番号や終了条件を自分で組み立てる余地を無くすための設計。
    2026-08-23 issue対応）。他セッションのプロセスや、ユーザーが対話的に
    使っている無関係なExcelウィンドウは、そもそもこのレジストリに記録され
    得ないため対象になり得ない。

    Returns:
        `[{"pid": int, "target_path": str, "terminated": bool}, ...]`。
        既に終了済みだったエントリはレジストリから除去した上で
        `terminated: True`として含める。終了に失敗したPIDはレジストリに
        残し`terminated: False`で報告する。
    """
    resolved = _excel_locks_registry()
    if resolved is None:
        return []
    registry_path, pm = resolved
    lock_path = registry_path.parent / f"{registry_path.name}.lock"
    with pm._locked(lock_path) as acquired:
        if not acquired:
            return []
        entries = _load_json_list(registry_path)
        results: list[dict] = []
        remaining: list[dict] = []
        for entry in entries:
            pid = entry.get("pid")
            target_path = entry.get("target_path")
            if not is_process_running(pid):
                results.append({"pid": pid, "target_path": target_path, "terminated": True})
                continue
            try:
                import win32api
                import win32con

                handle = win32api.OpenProcess(win32con.PROCESS_TERMINATE, False, pid)
                try:
                    win32api.TerminateProcess(handle, 1)
                finally:
                    win32api.CloseHandle(handle)
                wait_for_process_exit(pid, timeout_seconds=3.0)
                results.append({"pid": pid, "target_path": target_path, "terminated": True})
            except Exception:
                results.append({"pid": pid, "target_path": target_path, "terminated": False})
                remaining.append(entry)
        if remaining:
            _save_json_list(registry_path, remaining)
        elif registry_path.exists():
            registry_path.unlink()
    return results


def cell_to_json(value: object) -> object:
    """セル値をJSON化できる型へ変換する（日時系はISO8601文字列にする）。"""
    if isinstance(value, (datetime, date, time)):
        return value.isoformat()
    return value


def resolve_sheet_name(names: list[str], sheet_arg: str) -> str:
    """シート名の完全一致、次に0始まりインデックスとして解決する。"""
    if sheet_arg in names:
        return sheet_arg
    try:
        idx = int(sheet_arg)
    except ValueError:
        raise ValueError(f"シートが見つかりません: {sheet_arg}（存在するシート: {names}）")
    if 0 <= idx < len(names):
        return names[idx]
    raise ValueError(f"シートインデックスが範囲外です: {sheet_arg}（シート数: {len(names)}）")


def register_output_path(path, description: str | None = None) -> dict[str, str] | None:
    """生成/更新したファイルをパスメモリーへ登録し、{"@N": 絶対パス} を返す。

    run_script が子プロセスへ注入する AGENT_SRC_DIR 経由で src/path_memory.py
    を import する。AGENT_SRC_DIR未設定やimport失敗時はNone（run_script以外
    から直接実行された場合でもスクリプト自体は失敗させないためのフォールバック）。
    """
    src_dir = os.environ.get("AGENT_SRC_DIR")
    if not src_dir:
        return None
    if src_dir not in sys.path:
        sys.path.insert(0, src_dir)
    try:
        import path_memory
    except ImportError:
        return None
    thread_id, pm_dir, max_entries = path_memory.env_params()
    abs_path = str(Path(path).resolve())
    idx = path_memory.register(thread_id, abs_path, pm_dir, max_entries, description=description)
    if idx is None:
        return None
    return {f"@{idx}": abs_path}


def write_json_result(result: dict, category: str, source_path) -> dict:
    """result全体をJSONファイルへ書き出し、要約に足す {"result_path", "path_memory"} を返す。

    保存先は execute_python_code の中間生成物と同じ `_tmp_<thread_id>/<category>/`
    規約（pdf-tools の render_pdf_pages.py 参照）。基準は run_script の cwd
    ではなく常に default_workdir（AGENT_DEFAULT_WORKDIR）。cwd はユーザー指定
    work_dir になりうり、その場合は自動削除対象外のため、cwd基準だと消えずに
    溜まり続ける。
    """
    thread_id = os.environ.get("AGENT_EXEC_TMP_NAME") or os.environ.get("AGENT_THREAD_ID") or "_no_session"
    base_dir = Path(os.environ.get("AGENT_DEFAULT_WORKDIR") or "./data/temp")
    out_dir = base_dir / f"_tmp_{thread_id}" / category
    out_dir.mkdir(parents=True, exist_ok=True)
    digest = hashlib.sha1(str(Path(source_path).resolve()).encode("utf-8")).hexdigest()[:8]
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S_%f")
    out_path = out_dir / f"{digest}_{timestamp}.json"
    out_path.write_text(json.dumps(result, ensure_ascii=False, indent=2), encoding="utf-8")
    info: dict = {"result_path": str(out_path)}
    pm = register_output_path(out_path, description=f"{category}の読み込み結果全文JSON")
    if pm:
        info["path_memory"] = pm
    return info


def summarize_result(result: dict, omit_keys: list[str]) -> dict:
    """result から omit_keys で指定したキーを除いた要約辞書を返す。

    除いたキーがlist型なら "<key>_count"、str型なら "<key>_length" を
    件数/文字数として残す（本文全体はファイル側にのみ残す）。
    """
    summary = dict(result)
    for key in omit_keys:
        if key not in summary:
            continue
        value = summary.pop(key)
        if isinstance(value, list):
            summary[f"{key}_count"] = len(value)
        elif isinstance(value, str):
            summary[f"{key}_length"] = len(value)
    return summary


def _display_width(value: object) -> float:
    """全角文字（東アジアのWide/Fullwidth/Ambiguous）を半角2文字分として数えた表示幅。

    日本語混じりの表で `len(str(value))` を使うと全角文字が半角と同じ1として
    数えられ、実際の見た目より大幅に狭い列幅になってしまうため。
    """
    if value is None:
        return 0.0
    width = 0.0
    for ch in str(value):
        width += 2.0 if unicodedata.east_asian_width(ch) in ("W", "F", "A") else 1.0
    return width


def column_index(column: str) -> int:
    """列アルファベット（例"A"）または1始まりの列番号（数値文字列、例"3"）を列番号へ変換する。

    Excelの列番号はopenpyxl含め1始まりが一貫した規約のため（このスキル内の他のop、
    例えば`set_range`の`start_cell`も1始まりセル参照）、数値指定も1始まりとして扱う。
    0以下はopenpyxlのiter_rows等で列範囲指定が無視され意図しない全列走査になるため
    明示的にエラーにする。
    """
    from openpyxl.utils import column_index_from_string

    try:
        idx = column_index_from_string(str(column).upper())
    except ValueError:
        idx = int(column)
    if idx < 1:
        raise ValueError(f"列番号は1以上を指定してください（列アルファベットも使えます）: {column!r}")
    return idx


def group_column_values(ws, column: str, max_row: int | None = None) -> list[dict]:
    """指定列を上から走査し、非null値ごとに連続する行範囲をグループ化する。

    結合セルの非アンカー行（値はNone）、およびグルーピング目的で値を空欄に
    している行は、直前に出現した非null値のグループに属するとみなす
    （excel-edit SKILL.mdの「同じ値が続く列はセルを結合する」規約と対称）。

    Returns:
        [{"value": ..., "start_row": N, "end_row": N, "row_count": N}, ...]
        （行番号は1始まり。列内に一度もnull以外の値が現れない先頭行は
        どのグループにも属さず結果に含まれない）
    """
    col_idx = column_index(column)
    limit = max_row if max_row is not None else (ws.max_row or 0)
    groups: list[dict] = []
    current: dict | None = None
    if limit >= 1:
        for row_idx, (cell,) in enumerate(
            ws.iter_rows(min_row=1, max_row=limit, min_col=col_idx, max_col=col_idx),
            start=1,
        ):
            value = cell_to_json(cell.value)
            if value is not None:
                if current is not None:
                    groups.append(current)
                current = {"value": value, "start_row": row_idx, "end_row": row_idx, "row_count": 1}
            elif current is not None:
                current["end_row"] = row_idx
                current["row_count"] += 1
    if current is not None:
        groups.append(current)
    return groups
