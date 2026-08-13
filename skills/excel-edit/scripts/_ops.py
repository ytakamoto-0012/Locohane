"""edit_excel.py が適用する「操作（op）」のディスパッチ実装。

各関数は (workbook, op_dict) を受け取り、openpyxl のワークブックへ副作用を
適用する。値を返さない（呼び出し側が例外の有無だけを見る）。

run_script からは直接実行されない。edit_excel.py から import して使う。
"""

from __future__ import annotations

import sys
from pathlib import Path

# office_shared/excel_common.py から共有ヘルパーを import する（1-B 相互import方式）
_OFFICE_SHARED = Path(__file__).resolve().parent.parent.parent / "office_shared"
if str(_OFFICE_SHARED) not in sys.path:
    sys.path.append(str(_OFFICE_SHARED))
from excel_common import _display_width, column_index, group_column_values, resolve_sheet_name  # noqa: E402
from _style import apply_style, resolve_theme
from openpyxl.chart import BarChart, LineChart, PieChart, Reference, ScatterChart
from openpyxl.chart.label import DataLabelList
from openpyxl.formatting.rule import (
    CellIsRule,
    ColorScaleRule,
    DataBarRule,
    FormulaRule,
    IconSetRule,
)
from openpyxl.utils.cell import range_boundaries
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.worksheet.table import Table, TableStyleInfo

_CHART_CLASSES = {
    "bar": BarChart,
    "line": LineChart,
    "pie": PieChart,
    "scatter": ScatterChart,
}


def _sheet(wb, name: str):
    resolved = resolve_sheet_name(wb.sheetnames, name)
    return wb[resolved]


def _reference(ws, range_str: str) -> Reference:
    min_col, min_row, max_col, max_row = range_boundaries(range_str)
    return Reference(ws, min_col=min_col, min_row=min_row, max_col=max_col, max_row=max_row)


def _col_width_from_rows(rows: list[list[object]], start_col: int) -> dict[str, float]:
    from openpyxl.utils import get_column_letter

    widths: dict[str, float] = {}
    for row in rows:
        for offset, value in enumerate(row):
            letter = get_column_letter(start_col + offset)
            widths[letter] = max(widths.get(letter, 0.0), _display_width(value))
    return widths


def _apply_col_widths(ws, widths: dict[str, float]) -> None:
    """計算した列幅を反映する。既存の列幅より狭くなる場合は反映しない（縮めない）。

    同一シートへ `set_range`/`format_table` を複数回呼んでも、後から呼んだ
    範囲が狭いという理由で先に確保した幅が失われないようにするため。
    """
    for letter, width in widths.items():
        target = min(max(width + 2, 8), 60)
        current = ws.column_dimensions[letter].width
        if current is None or target > current:
            ws.column_dimensions[letter].width = target


def op_add_sheet(wb, op: dict) -> None:
    name = str(op["name"])[:31]
    if name in wb.sheetnames:
        raise ValueError(f"シート'{name}'は既に存在します（rename_sheetで別名にするか、" "既存シートへの書き込みに切り替えてください）")
    wb.create_sheet(title=name, index=op.get("index"))


def op_delete_sheet(wb, op: dict) -> None:
    wb.remove(_sheet(wb, op["name"]))


def op_rename_sheet(wb, op: dict) -> None:
    _sheet(wb, op["name"]).title = str(op["new_name"])[:31]


def op_set_cell(wb, op: dict) -> None:
    ws = _sheet(wb, op["sheet"])
    cell = ws[op["cell"]]
    cell.value = op.get("value")
    apply_style(cell, op.get("style"))


def _write_rows(ws, min_row: int, min_col: int, op: dict) -> None:
    """`rows`（+ style/header_style/row_styles/format_table）を起点セルから書き込む。

    `op_set_range`（起点セル明示）と`op_insert_row_group`（挿入位置解決後）の
    両方から呼ばれる共通実装。
    """
    rows = op.get("rows") or []
    style = op.get("style")
    header_style = op.get("header_style", style)
    row_styles = op.get("row_styles")
    if row_styles is not None and len(row_styles) != len(rows):
        raise ValueError(f"row_stylesの要素数({len(row_styles)})はrowsの行数({len(rows)})と" "一致させてください（見出し行分も含めて1要素=1行）")
    for r_offset, row in enumerate(rows):
        if row_styles is not None:
            row_style = row_styles[r_offset]
        else:
            row_style = header_style if r_offset == 0 else style
        for c_offset, value in enumerate(row):
            cell = ws.cell(row=min_row + r_offset, column=min_col + c_offset, value=value)
            apply_style(cell, row_style)
    _apply_col_widths(ws, _col_width_from_rows(rows, min_col))

    format_table_opt = op.get("format_table")
    # header_style を渡した時点で「1行目は見出し」という意図が明示されているので、
    # 明示的に format_table=false でオプトアウトしない限り既定で美しい表に仕上げる。
    # LLM（特に指示追従力の弱い小型ローカルモデル）に新しいフラグを覚えさせず、
    # 既存の呼び出しパターン（header_styleを渡す）だけで綺麗な見た目になるようにするため。
    should_format = format_table_opt is not False and (format_table_opt or "header_style" in op)
    if should_format and rows:
        max_col = min_col + max(len(row) for row in rows) - 1
        max_row = min_row + len(rows) - 1
        opts = format_table_opt if isinstance(format_table_opt, dict) else {}
        _format_table_range(ws, min_row, max_row, min_col, max_col, opts)


def op_set_range(wb, op: dict) -> None:
    ws = _sheet(wb, op["sheet"])
    min_col, min_row, _, _ = range_boundaries(f"{op['start_cell']}:{op['start_cell']}")
    _write_rows(ws, min_row, min_col, op)


def op_insert_row_group(wb, op: dict) -> None:
    """列の値（アンカー）を基準に、その場で挿入位置を解決してから挿入・書き込みまでを
    1opで完結させる。`insert_rows`+`set_range`を絶対行番号で組み合わせる方式と異なり、
    このop実行時点（＝同一opsバッチ内の直前のopsが既に適用済みの最新のシート状態）から
    アンカーを解決するため、複数の`insert_row_group`を連続実行しても行ズレによる
    上書き事故が起きない。
    """
    ws = _sheet(wb, op["sheet"])
    rows = op.get("rows") or []
    anchor = op.get("anchor")
    if anchor:
        matches = [g for g in group_column_values(ws, anchor["column"]) if g["value"] == anchor["equals"]]
        if not matches:
            raise ValueError(
                "insert_row_groupのanchorに一致する値が見つかりません: " f"column={anchor.get('column')!r} equals={anchor.get('equals')!r}"
            )
        target = matches[0]
        position = op.get("position", "before")
        if position == "before":
            index = target["start_row"]
        elif position == "after":
            index = target["end_row"] + 1
        else:
            raise ValueError(f"insert_row_groupのpositionは'before'/'after'のみ対応です: {position!r}")
        min_col = column_index(op["start_column"]) if op.get("start_column") else column_index(anchor["column"])
    else:
        index = (ws.max_row or 0) + 1
        min_col = column_index(op["start_column"]) if op.get("start_column") else 1

    _resize_preserving_merges(ws, "row", index, len(rows))
    _write_rows(ws, index, min_col, op)


def op_set_style(wb, op: dict) -> None:
    ws = _sheet(wb, op["sheet"])
    # ws[range] は単一セル指定（例"A1"）だとCellを、複数セル範囲だと
    # タプルのタプルを返し戻り値の型が揺れるため、range_boundariesで
    # 座標に正規化してcell()で個別に取得する（set_rangeと同じ方式）。
    min_col, min_row, max_col, max_row = range_boundaries(op["range"])
    for row in range(min_row, max_row + 1):
        for col in range(min_col, max_col + 1):
            apply_style(ws.cell(row=row, column=col), op.get("style"))


def _format_table_range(ws, min_row: int, max_row: int, min_col: int, max_col: int, opts: dict) -> None:
    """指定範囲（1行目=見出し行）へ、配色・罫線・縞模様・見出し固定・列幅調整を
    一括で適用する。`op_format_table`（独立op）と `op_set_range` の
    `format_table` インラインオプションの両方から使う共通実装。

    本体行のスタイルには意図的に font 関連キーを含めない
    （`apply_style`はfont系キーが無ければcell.fontに触れないため、`role`規約等で
    既に当ててある文字色・太字を上書きしない）。
    """
    border_style = opts.get("border", "thin")
    banded = opts.get("banded", True)
    band_fill = opts.get("band_fill", "F2F2F2")
    theme = resolve_theme(opts["theme"]) if opts.get("theme") else None
    header_fill = opts.get("header_fill") or (theme["primary"] if theme else "1F4E78")
    header_font_color_default = theme["text_on_primary"] if theme else "FFFFFF"

    # min_row の直前行が既に見出し色（header_fill）で塗られている場合、この
    # 呼び出しは新規テーブルの先頭ではなく、見出し行とデータ行を別々の
    # set_range 呼び出しに分けてしまった「既存テーブルへの追記」とみなし、
    # min_row を見出しとして再装飾しない。LLMが `header_style` を見出し行と
    # データ行の両方の呼び出しに付けてしまっても、データ側の先頭行が誤って
    # 見出し配色になる事故を防ぐため。
    prev_fill_rgb = str(ws.cell(row=min_row - 1, column=min_col).fill.fgColor.rgb or "") if min_row > 1 else ""
    is_continuation = prev_fill_rgb.upper().endswith(header_fill.upper())

    body_start = min_row
    if not is_continuation:
        header_style = {
            "bold": True,
            "font_color": opts.get("header_font_color") or header_font_color_default,
            "fill_color": header_fill,
            "align": "center",
            "valign": "center",
            "border": border_style,
        }
        for col in range(min_col, max_col + 1):
            apply_style(ws.cell(row=min_row, column=col), header_style)
        body_start = min_row + 1

    for row_idx, row in enumerate(range(body_start, max_row + 1)):
        body_style = {"border": border_style}
        if banded and row_idx % 2 == 1:
            body_style["fill_color"] = band_fill
        for col in range(min_col, max_col + 1):
            apply_style(ws.cell(row=row, column=col), body_style)

    if opts.get("freeze_header", True):
        ws.freeze_panes = ws.cell(row=body_start, column=min_col).coordinate

    if opts.get("autofit", True):
        rows_values = [[ws.cell(row=r, column=c).value for c in range(min_col, max_col + 1)] for r in range(min_row, max_row + 1)]
        _apply_col_widths(ws, _col_width_from_rows(rows_values, min_col))


def op_format_table(wb, op: dict) -> None:
    ws = _sheet(wb, op["sheet"])
    min_col, min_row, max_col, max_row = range_boundaries(op["range"])
    _format_table_range(ws, min_row, max_row, min_col, max_col, op)


def _shift_index(value: int, index: int, amount: int) -> int:
    """行/列インデックスvalueを、amount>0なら「index位置の直前にamount本挿入」、
    amount<0なら「index位置からabs(amount)本削除」した後の新しい位置へ変換する。

    削除で消える位置（index <= value < index-amount）はindexへクランプする。
    区間の上端（hi）をこの関数だけで変換すると「削除範囲の内側で始まり外側で
    終わる」区間の長さがズレるため、呼び出し側（`_merged_ranges_after_resize`）
    では hi は区間長を引き算し直して求める。
    """
    if amount > 0:
        return value if value < index else value + amount
    count = -amount
    if value < index:
        return value
    if value < index + count:
        return index
    return value - count


def _merged_ranges_after_resize(saved: list[tuple[int, int, int, int]], axis: str, index: int, amount: int):
    """`saved`（挿入・削除前の結合範囲の(min_col,min_row,max_col,max_row)一覧）を、
    挿入・削除後の座標へ変換して返す。復元不能（削除範囲に完全に収まっていた・
    1x1セルに縮んだ）な範囲は結果から除外する。

    amount>0: index位置の直前にamount本を挿入。amount<0: index位置からabs(amount)本を削除。
    """
    is_row = axis == "row"
    for min_col, min_row, max_col, max_row in saved:
        lo, hi = (min_row, max_row) if is_row else (min_col, max_col)
        length = hi - lo + 1

        if amount > 0:
            new_lo = _shift_index(lo, index, amount)
            new_hi = _shift_index(hi, index, amount)
        else:
            count = -amount
            overlap = max(0, min(hi, index + count - 1) - max(lo, index) + 1)
            if overlap >= length:
                continue  # 削除範囲に完全に収まっていた結合 → 復元不能なので破棄する
            new_lo = _shift_index(lo, index, amount)
            new_hi = new_lo + (length - overlap) - 1

        if is_row:
            new_min_row, new_max_row, new_min_col, new_max_col = new_lo, new_hi, min_col, max_col
        else:
            new_min_row, new_max_row, new_min_col, new_max_col = min_row, max_row, new_lo, new_hi

        if new_min_row == new_max_row and new_min_col == new_max_col:
            continue  # 縦横とも1セルにまで縮んだ場合は結合不要
        yield (new_min_col, new_min_row, new_max_col, new_max_row)


def _resize_preserving_merges(ws, axis: str, index: int, amount: int) -> None:
    """insert_rows/delete_rows/insert_cols/delete_cols/insert_row_group の
    挿入・削除処理をラップし、既存の結合セル範囲を正しく追従させる。

    openpyxlのinsert_rows等はセルの値だけを移動し、既存の結合セル範囲
    （`ws.merged_cells.ranges`）の座標も、各セルの実際の型（`MergedCell`/`Cell`）の
    対応関係も追従させない（openpyxl 3.1系の既知の制限。`Worksheet._move_cells`は
    `self._cells`しか動かさない）。挿入・削除の"あと"に結合範囲の座標だけを
    机上で計算し直して`merge_cells`/`unmerge_cells`し直すアプローチは、その時点で
    実際のセル（`MergedCell`）が既に黙って移動済みのため`KeyError`等で破綻する。

    そのため、挿入・削除の"前"に一旦すべての結合を解除してアンカーセルの値・
    書式だけが残る単純な配置に戻し、openpyxl本体の`insert_rows`等へ素直に
    移動を任せてから、新しい座標を計算して結合をかけ直す。

    amount>0: index位置の直前にamount本を挿入する。
    amount<0: index位置からabs(amount)本を削除する。
    """
    if amount == 0:
        return

    saved = []
    for merged_range in list(ws.merged_cells.ranges):
        min_col, min_row, max_col, max_row = merged_range.bounds
        saved.append((min_col, min_row, max_col, max_row))
        ws.unmerge_cells(start_row=min_row, start_column=min_col, end_row=max_row, end_column=max_col)

    if axis == "row":
        ws.insert_rows(index, amount) if amount > 0 else ws.delete_rows(index, -amount)
    else:
        ws.insert_cols(index, amount) if amount > 0 else ws.delete_cols(index, -amount)

    for min_col, min_row, max_col, max_row in _merged_ranges_after_resize(saved, axis, index, amount):
        ws.merge_cells(start_row=min_row, start_column=min_col, end_row=max_row, end_column=max_col)


def op_insert_rows(wb, op: dict) -> None:
    _resize_preserving_merges(_sheet(wb, op["sheet"]), "row", op["index"], op.get("count", 1))


def op_delete_rows(wb, op: dict) -> None:
    _resize_preserving_merges(_sheet(wb, op["sheet"]), "row", op["index"], -op.get("count", 1))


def op_insert_cols(wb, op: dict) -> None:
    _resize_preserving_merges(_sheet(wb, op["sheet"]), "col", op["index"], op.get("count", 1))


def op_delete_cols(wb, op: dict) -> None:
    _resize_preserving_merges(_sheet(wb, op["sheet"]), "col", op["index"], -op.get("count", 1))


def op_set_column_width(wb, op: dict) -> None:
    width = min(max(float(op["width"]), 1), 60)
    _sheet(wb, op["sheet"]).column_dimensions[op["column"]].width = width


def op_set_row_height(wb, op: dict) -> None:
    height = min(max(float(op["height"]), 1), 409)
    _sheet(wb, op["sheet"]).row_dimensions[int(op["row"])].height = height


def op_merge_cells(wb, op: dict) -> None:
    ws = _sheet(wb, op["sheet"])
    min_col, min_row, max_col, max_row = range_boundaries(op["range"])
    conflicts = [
        str(r)
        for r in ws.merged_cells.ranges
        if not (max_col < r.min_col or min_col > r.max_col or max_row < r.min_row or min_row > r.max_row)
    ]
    if conflicts:
        raise ValueError(
            f"merge_cellsの範囲 {op['range']!r} は既存の結合範囲 {conflicts} と重複しています。"
            "先にunmerge_cellsで解除するか、範囲を見直してください。"
        )
    ws.merge_cells(op["range"])


def op_unmerge_cells(wb, op: dict) -> None:
    _sheet(wb, op["sheet"]).unmerge_cells(op["range"])


def op_add_table(wb, op: dict) -> None:
    ws = _sheet(wb, op["sheet"])
    tbl = Table(displayName=op["name"], ref=op["range"])
    tbl.tableStyleInfo = TableStyleInfo(
        name=op.get("style", "TableStyleMedium9"),
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=op.get("banded", True),
        showColumnStripes=False,
    )
    ws.add_table(tbl)


def op_update_table(wb, op: dict) -> None:
    ws = _sheet(wb, op["sheet"])
    tbl = ws.tables.get(op["name"])
    if tbl is None:
        raise KeyError(f"テーブルが見つかりません: {op['name']!r}")
    if op.get("range"):
        tbl.ref = op["range"]
    if op.get("style"):
        tbl.tableStyleInfo.name = op["style"]
    if "banded" in op:
        tbl.tableStyleInfo.showRowStripes = bool(op["banded"])


def op_remove_table(wb, op: dict) -> None:
    ws = _sheet(wb, op["sheet"])
    if op["name"] not in ws.tables:
        raise KeyError(f"テーブルが見つかりません: {op['name']!r}")
    del ws.tables[op["name"]]


def op_freeze_panes(wb, op: dict) -> None:
    _sheet(wb, op["sheet"]).freeze_panes = op["cell"]


def op_add_chart(wb, op: dict) -> None:
    ws = _sheet(wb, op["sheet"])
    chart_cls = _CHART_CLASSES.get(op["type"])
    if chart_cls is None:
        raise ValueError(f"未対応のグラフ種別です（bar/line/pie/scatterのみ対応）: {op['type']}")
    chart = chart_cls()
    if op.get("title"):
        chart.title = op["title"]
    data_ref = _reference(ws, op["data_range"])
    chart.add_data(data_ref, titles_from_data=op.get("titles_from_data", True))
    if op.get("categories_range"):
        chart.set_categories(_reference(ws, op["categories_range"]))

    # 既定でデータラベルを表示する（公式pptxスキルが指摘する
    # 「装飾なしのグラフは見劣りする」問題への対応。無地の棒/線グラフより
    # 値が読み取れるほうが実用上有用なため、明示的にオフにしない限り常時表示）。
    if op.get("show_data_labels", True):
        chart.dataLabels = DataLabelList(
            showVal=False,
            showCatName=False,
            showSerName=False,
            showLegendKey=False,
            showPercent=False,
            showBubbleSize=False,
        )
        if op["type"] == "pie":
            chart.dataLabels.showPercent = True
        else:
            chart.dataLabels.showVal = True

    theme = resolve_theme(op["theme"]) if op.get("theme") else None
    if theme is not None and op["type"] != "pie":
        palette = [theme["primary"], theme["secondary"], theme["accent"]]
        for i, series in enumerate(chart.series):
            series.graphicalProperties.solidFill = palette[i % len(palette)]

    ws.add_chart(chart, op["anchor"])


_CF_BUILDERS = {
    "color_scale": lambda p: ColorScaleRule(**p),
    "cell_is": lambda p: CellIsRule(**p),
    "formula": lambda p: FormulaRule(**p),
    "data_bar": lambda p: DataBarRule(**p),
    "icon_set": lambda p: IconSetRule(**p),
}


def op_add_conditional_format(wb, op: dict) -> None:
    ws = _sheet(wb, op["sheet"])
    builder = _CF_BUILDERS.get(op["rule_type"])
    if builder is None:
        raise ValueError(f"未対応のrule_typeです（color_scale/cell_is/formula/data_bar/icon_setのみ対応）: {op['rule_type']}")
    try:
        rule = builder(op.get("params") or {})
    except TypeError as e:
        raise ValueError(f"params が rule_type={op['rule_type']!r} に対して不正です: {e}")
    ws.conditional_formatting.add(op["range"], rule)


def op_add_data_validation(wb, op: dict) -> None:
    ws = _sheet(wb, op["sheet"])
    dv = DataValidation(
        type=op["type"],
        formula1=op.get("formula1"),
        formula2=op.get("formula2"),
        allow_blank=op.get("allow_blank", True),
    )
    if op.get("prompt"):
        dv.prompt = op["prompt"]
        dv.promptTitle = op.get("prompt_title", "")
        dv.showInputMessage = True
    if op.get("error_message"):
        dv.error = op["error_message"]
        dv.errorTitle = op.get("error_title", "")
        dv.showErrorMessage = True
    ws.add_data_validation(dv)
    dv.add(op["range"])


OP_HANDLERS = {
    "add_sheet": op_add_sheet,
    "delete_sheet": op_delete_sheet,
    "rename_sheet": op_rename_sheet,
    "set_cell": op_set_cell,
    "set_range": op_set_range,
    "set_style": op_set_style,
    "format_table": op_format_table,
    "insert_rows": op_insert_rows,
    "insert_row_group": op_insert_row_group,
    "delete_rows": op_delete_rows,
    "insert_cols": op_insert_cols,
    "delete_cols": op_delete_cols,
    "set_column_width": op_set_column_width,
    "set_row_height": op_set_row_height,
    "merge_cells": op_merge_cells,
    "unmerge_cells": op_unmerge_cells,
    "add_table": op_add_table,
    "update_table": op_update_table,
    "remove_table": op_remove_table,
    "freeze_panes": op_freeze_panes,
    "add_chart": op_add_chart,
    "add_conditional_format": op_add_conditional_format,
    "add_data_validation": op_add_data_validation,
}


def apply_op(wb, op: dict) -> None:
    op_name = op.get("op")
    handler = OP_HANDLERS.get(op_name)
    if handler is None:
        raise ValueError(f"未対応のopです: {op_name!r}（対応op: {sorted(OP_HANDLERS)}）")
    handler(wb, op)
