"""excel-edit・excel-read 両スキルが共有する列グルーピングロジック。

実体はこのファイル（excel-editスキル側）のみ。excel-read側は複製を持たず、
sys.path経由で本ファイルを直接importする（read_excel.py参照）。理由は
「同じロジックを2箇所に別々実装すると、読み込み側の報告と書き込み側の
挿入位置解決が食い違い、新たな不整合を生む」ため（詳細はissue参照）。

run_script からは直接実行されない。edit_excel.py（_ops.py経由）・
read_excel.py から import して使う。
"""

from __future__ import annotations

from _common import cell_to_json


def column_index(column: str) -> int:
    """列アルファベット（例"A"）または1始まりの列番号（数値文字列、例"3"）を列番号へ変換する。

    Excelの列番号はopenpyxl含め1始まりが一貫した規約のため（このスキル内の他のop、
    例えば`set_range`の`start_cell`も1始まりセル参照）、数値指定も1始まりとして扱う。
    0以下はopenpyxlのiter_rows等で列範囲指定が無視され意図しない全列走査になるため
    明示的にエラーにする。
    """
    from openpyxl.utils import column_index_from_string

    try:
        idx = column_index_from_string(str(column).upper())
    except ValueError:
        idx = int(column)
    if idx < 1:
        raise ValueError(f"列番号は1以上を指定してください（列アルファベットも使えます）: {column!r}")
    return idx


def group_column_values(ws, column: str, max_row: int | None = None) -> list[dict]:
    """指定列を上から走査し、非null値ごとに連続する行範囲をグループ化する。

    結合セルの非アンカー行（値はNone）、およびグルーピング目的で値を空欄に
    している行は、直前に出現した非null値のグループに属するとみなす
    （excel-edit SKILL.mdの「同じ値が続く列はセルを結合する」規約と対称）。

    Returns:
        [{"value": ..., "start_row": N, "end_row": N, "row_count": N}, ...]
        （行番号は1始まり。列内に一度もnull以外の値が現れない先頭行は
        どのグループにも属さず結果に含まれない）
    """
    col_idx = column_index(column)
    limit = max_row if max_row is not None else (ws.max_row or 0)
    groups: list[dict] = []
    current: dict | None = None
    if limit >= 1:
        for row_idx, (cell,) in enumerate(
            ws.iter_rows(min_row=1, max_row=limit, min_col=col_idx, max_col=col_idx),
            start=1,
        ):
            value = cell_to_json(cell.value)
            if value is not None:
                if current is not None:
                    groups.append(current)
                current = {"value": value, "start_row": row_idx, "end_row": row_idx, "row_count": 1}
            elif current is not None:
                current["end_row"] = row_idx
                current["row_count"] += 1
    if current is not None:
        groups.append(current)
    return groups
