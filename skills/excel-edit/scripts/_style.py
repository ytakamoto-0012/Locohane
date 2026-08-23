"""style辞書（JSON）と openpyxl のスタイルオブジェクトを相互変換するヘルパー。

同スキル内の edit_excel.py（書き込み: apply_style）から import して使う。
excel-read/scripts/_style.py は同名だが独立した別ファイル（extract_style用。
office_theme.pyに依存しない点がこちらと異なる）。変更してもこちらには反映されない。
run_script からは直接実行されない。

style辞書の形式（すべて省略可）:
{
  "bold": true, "italic": false,
  "font_color": "0000FF", "font_size": 11,
  "fill_color": "FFFF00",
  "number_format": "#,##0.00",
  "align": "center", "valign": "center", "wrap_text": false,
  "border": "thin",                      # または {"top": "thin", "bottom": "thin", ...}
  "role": "input"                        # font_color省略時のショートカット
}

role のショートカット（Anthropic公式スキルの財務モデル色分け規約に準拠）:
- "input"   -> 青 0000FF（ハードコードされた入力値）
- "formula" -> 黒 000000（同一シート内の数式）
- "link"    -> 緑 008000（他シートからの参照）
"""

from __future__ import annotations

import sys
from pathlib import Path

from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

ROLE_COLORS = {
    "input": "0000FF",
    "formula": "000000",
    "link": "008000",
}

# office_shared/office_theme.py から THEMES / resolve_theme を import する（1-B 相互import方式）
_OFFICE_SHARED = Path(__file__).resolve().parent.parent.parent / "office_shared"
if str(_OFFICE_SHARED) not in sys.path:
    sys.path.append(str(_OFFICE_SHARED))
from office_theme import THEMES, resolve_theme  # noqa: E402

_BORDER_SIDES = ("top", "bottom", "left", "right")


def build_font(style: dict) -> Font | None:
    if not any(k in style for k in ("bold", "italic", "font_color", "font_size", "role")):
        return None
    color = style.get("font_color") or ROLE_COLORS.get(style.get("role", ""))
    return Font(
        bold=bool(style.get("bold", False)),
        italic=bool(style.get("italic", False)),
        color=color,
        size=style.get("font_size"),
    )


def build_fill(style: dict) -> PatternFill | None:
    color = style.get("fill_color")
    if not color:
        return None
    return PatternFill(fill_type="solid", start_color=color, end_color=color)


def build_alignment(style: dict) -> Alignment | None:
    if not any(k in style for k in ("align", "valign", "wrap_text")):
        return None
    return Alignment(
        horizontal=style.get("align"),
        vertical=style.get("valign"),
        wrap_text=bool(style.get("wrap_text", False)),
    )


def build_border(style: dict) -> Border | None:
    spec = style.get("border")
    if not spec:
        return None
    if isinstance(spec, str):
        sides = {side: spec for side in _BORDER_SIDES}
    elif isinstance(spec, dict):
        sides = spec
    else:
        raise ValueError(f"border の指定が不正です: {spec!r}")
    return Border(**{side: Side(style=sides[side]) for side in sides if sides.get(side)})


def apply_style(cell, style: dict | None) -> None:
    """1セルに style 辞書の内容を適用する。"""
    if not style:
        return
    font = build_font(style)
    if font is not None:
        cell.font = font
    fill = build_fill(style)
    if fill is not None:
        cell.fill = fill
    alignment = build_alignment(style)
    if alignment is not None:
        cell.alignment = alignment
    border = build_border(style)
    if border is not None:
        cell.border = border
    if style.get("number_format"):
        cell.number_format = style["number_format"]


def _normalize_rgb(color) -> str | None:
    """openpyxlのColorオブジェクトからRRGGBB文字列を取り出す。

    ARGB（例 "00FFFF00"）は先頭2桁のアルファ成分を落とす。テーマ色や
    インデックスカラー等、単純なrgb文字列でない場合はNoneを返す。
    """
    if color is None:
        return None
    rgb = getattr(color, "rgb", None)
    if not isinstance(rgb, str):
        return None
    return rgb[-6:] if len(rgb) == 8 else rgb


def extract_style(cell) -> dict:
    """1セルの書式を、apply_style()が受け取るstyle辞書と同じキー体系で返す。

    既定値と一致するプロパティは省略し、実際に設定されている項目だけを返す。
    """
    style: dict = {}

    font = cell.font
    if font is not None:
        if font.bold:
            style["bold"] = True
        if font.italic:
            style["italic"] = True
        font_color = _normalize_rgb(font.color)
        if font_color and font_color != "000000":
            style["font_color"] = font_color
        if font.size and font.size != 11:
            style["font_size"] = font.size

    fill = cell.fill
    if fill is not None and fill.fill_type == "solid":
        fill_color = _normalize_rgb(fill.fgColor)
        if fill_color and fill_color != "FFFFFF":
            style["fill_color"] = fill_color

    alignment = cell.alignment
    if alignment is not None:
        if alignment.horizontal:
            style["align"] = alignment.horizontal
        if alignment.vertical:
            style["valign"] = alignment.vertical
        if alignment.wrap_text:
            style["wrap_text"] = True

    border = cell.border
    if border is not None:
        sides = {side: getattr(border, side).style for side in _BORDER_SIDES}
        sides = {side: v for side, v in sides.items() if v}
        if sides:
            values = set(sides.values())
            style["border"] = next(iter(values)) if len(values) == 1 and len(sides) == 4 else sides

    if cell.number_format and cell.number_format != "General":
        style["number_format"] = cell.number_format

    return style
