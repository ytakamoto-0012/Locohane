"""ops（操作）のリストを適用して xlsx/xlsm を新規作成または編集する。

excel-tools スキルの実行スクリプト（progressive disclosure 第3段階）。
run_script から
    python edit_excel.py <path> --ops-json "<JSON配列>" [--new] [--overwrite] [--output <別名保存先>]
または
    python edit_excel.py <path> --ops-file <JSONファイルのパス> [--new] [--overwrite] [--output ...]
の形で呼ばれる（このプロジェクトには汎用のファイル書き込みツールが無いため、
LLMが組み立てたJSONをそのまま --ops-json 引数として渡せるようにしている）。

--new を付けると新規ワークブックを作成する（既存ファイルがある場合は
--overwrite 必須）。--new を付けない場合は <path> を読み込んで編集する
（存在しなければエラー）。--output を省略した場合は <path> へ上書き保存する。

ops の各要素の形式や対応opの一覧は SKILL.md を参照。実装（各opの処理）は
scripts/_ops.py の OP_HANDLERS を参照。
"""

from __future__ import annotations

import argparse
import json
import sys
from pathlib import Path

from _common import backup_before_overwrite, register_output_path, setup_utf8_stdio
from _ops import apply_op


def _load_json_arg(args: argparse.Namespace) -> str:
    if args.ops_json is not None:
        return args.ops_json
    data_path = Path(args.ops_file)
    if not data_path.is_file():
        raise FileNotFoundError(f"opsファイルが見つかりません: {args.ops_file}")
    try:
        return data_path.read_text(encoding="utf-8")
    except UnicodeDecodeError:
        return data_path.read_text(encoding="cp932")


def main() -> int:
    setup_utf8_stdio()
    parser = argparse.ArgumentParser()
    parser.add_argument("path", help="編集対象（--new時は作成先）のxlsx/xlsmパス")
    group = parser.add_mutually_exclusive_group(required=True)
    group.add_argument("--ops-json", help="適用する操作のJSON配列をそのまま渡す")
    group.add_argument("--ops-file", help="適用する操作のJSON配列を書いたファイルのパス")
    parser.add_argument("--new", action="store_true", help="既存ファイルを読まず新規ワークブックとして作成する")
    parser.add_argument("--overwrite", action="store_true", help="--new時に出力先が既に存在する場合の上書きを許可する")
    parser.add_argument("--output", default=None, help="保存先パス（省略時は path へ上書き保存）")
    args = parser.parse_args()

    path = Path(args.path)
    ext = path.suffix.lower()
    if ext not in (".xlsx", ".xlsm"):
        print(f"対応拡張子は .xlsx/.xlsm のみです: {ext}", file=sys.stderr)
        return 1

    output_path = Path(args.output) if args.output else path

    try:
        raw = _load_json_arg(args)
    except FileNotFoundError as e:
        print(str(e), file=sys.stderr)
        return 1

    try:
        ops = json.loads(raw)
    except json.JSONDecodeError as e:
        print(f"opsのJSON解析に失敗しました: {e}", file=sys.stderr)
        return 1
    if not isinstance(ops, list):
        print("opsはJSON配列である必要があります", file=sys.stderr)
        return 1

    import openpyxl

    if args.new:
        if path.exists() and not args.overwrite:
            print(f"作成先が既に存在します（上書きするには --overwrite を指定）: {args.path}", file=sys.stderr)
            return 1
        wb = openpyxl.Workbook()
        wb.remove(wb.active)
    else:
        if not path.is_file():
            print(f"編集対象のファイルが見つかりません（新規作成する場合は --new を指定）: {args.path}", file=sys.stderr)
            return 1
        try:
            wb = openpyxl.load_workbook(str(path), keep_vba=(ext == ".xlsm"))
        except Exception as e:  # noqa: BLE001 - 破損ファイル等、openpyxlが送出する多様な例外を丸めて報告する
            print(f"ファイルの読み込みに失敗しました: {e}", file=sys.stderr)
            return 1

    for idx, op in enumerate(ops):
        if not isinstance(op, dict) or "op" not in op:
            print(f"ops[{idx}] が不正です（'op'キーを持つオブジェクトである必要があります）: {op!r}", file=sys.stderr)
            return 1
        try:
            apply_op(wb, op)
        except (KeyError, ValueError, TypeError) as e:
            print(f"ops[{idx}]（op={op.get('op')!r}）の適用に失敗しました: {e}", file=sys.stderr)
            return 1

    if not wb.sheetnames:
        print("最終的にシートが1つも存在しません（add_sheetを含めるか--newを見直してください）", file=sys.stderr)
        return 1

    output_path.parent.mkdir(parents=True, exist_ok=True)
    backup_path = backup_before_overwrite(output_path)
    try:
        wb.save(str(output_path))
    except OSError as e:
        print(f"ファイルの保存に失敗しました: {e}", file=sys.stderr)
        return 1

    result = {
        "path": str(output_path),
        "backup_path": str(backup_path) if backup_path else None,
        "sheets": wb.sheetnames,
        "applied_ops": len(ops),
    }
    path_memory = register_output_path(output_path, description="edit_excelが生成/更新したファイル")
    if path_memory:
        result["path_memory"] = path_memory
    print(json.dumps(result, ensure_ascii=False))
    return 0


if __name__ == "__main__":
    sys.exit(main())
