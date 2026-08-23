"""xlsx/xlsm/xls ファイルのシート一覧またはセルデータを読み込みJSONで出力する。

excel-tools スキルの実行スクリプト（progressive disclosure 第3段階）。
run_script から
    python read_excel.py <file_path> [--sheet <シート名またはインデックス>]
                          [--offset N] [--limit N] [--data-only]
の形で呼ばれる。

拡張子でライブラリを切り替える:
- .xlsx / .xlsm -> openpyxl（style取得のため常に通常モードで読む）
- .xls          -> xlrd（xlrd 2.x はレガシー .xls 専用。.xlsx は読めない。
                    styleは非対応で常に値のみ）

--sheet を省略した場合はシート名と行数・列数の一覧のみを返す（大きいファイルを
うっかり全件読み込まないための既定動作）。--sheet を指定した場合のみセル
データ本体を --offset/--limit の範囲で返す。

xlsx/xlsmでは--sheet指定時、太字・背景色・セル結合・構造化テーブルなどの
style情報も常に合わせて返す（低パラメータモデルでもstyle情報を見落とさず
不具合判断できるようにするため）。セル単位の書式（apply_styleと同じキー体系）
に加え、シート全体のmerged_cells/tablesも返す。openpyxlのread_onlyモードは
使わないため、通常モードよりメモリ効率は落ちる。
"""

from __future__ import annotations

import argparse
import json
import sys
import traceback
from pathlib import Path

from _style import extract_style

# office_shared/excel_common.py から共有ヘルパーを import する（1-B 相互import方式）。
# group_by等の列グルーピングロジックは読み書き双方で同一実装を使う必要があり
# （複製すると読み込み結果と書き込み側の挿入位置解決が食い違う恐れがある）、
# _common系のヘルパーと合わせてexcel_common.pyに一本化している。
_OFFICE_SHARED = Path(__file__).resolve().parent.parent.parent / "office_shared"
if str(_OFFICE_SHARED) not in sys.path:
    sys.path.append(str(_OFFICE_SHARED))
from excel_common import (  # noqa: E402
    _display_width,
    cell_to_json,
    group_column_values,
    resolve_sheet_name,
    setup_utf8_stdio,
    summarize_result,
    write_json_result,
)
from openpyxl.drawing.spreadsheet_drawing import AbsoluteAnchor
from openpyxl.utils.units import EMU_to_cm


import re

# セル参照（"$"付き列/行を許容、シート修飾された参照は除外して誤検知を防ぐ）。
# 関数名（SUM等）は数字が直後に続かないためマッチしない。
_CELL_RANGE_RE = re.compile(
    r"(?<![!$A-Za-z0-9_])(\$?[A-Za-z]{1,3})(\$?[0-9]+)(?::(\$?[A-Za-z]{1,3})(\$?[0-9]+))?"
)


def _formula_self_references(formula: str, own_col: int, own_row: int) -> bool:
    """formula中のセル参照（自シート内、シート修飾なしのもののみ）が
    own_col/own_row自身を範囲に含むかを判定する。

    シート修飾された参照（Sheet2!A1等）は判定対象外にして誤検知を避ける
    （正規表現の直前が"!"の参照は`_CELL_RANGE_RE`が最初からマッチしない）。
    """
    from openpyxl.utils import column_index_from_string

    for match in _CELL_RANGE_RE.finditer(formula):
        c1_letters, r1_digits, c2_letters, r2_digits = match.groups()
        try:
            col1 = column_index_from_string(c1_letters.lstrip("$"))
            row1 = int(r1_digits.lstrip("$"))
        except ValueError:
            continue
        if c2_letters is not None and r2_digits is not None:
            try:
                col2 = column_index_from_string(c2_letters.lstrip("$"))
                row2 = int(r2_digits.lstrip("$"))
            except ValueError:
                continue
            min_col, max_col = min(col1, col2), max(col1, col2)
            min_row, max_row = min(row1, row2), max(row1, row2)
        else:
            min_col = max_col = col1
            min_row = max_row = row1
        if min_col <= own_col <= max_col and min_row <= own_row <= max_row:
            return True
    return False


def _query_group_by(ws, query: dict, max_row: int) -> dict:
    if "column" not in query:
        raise ValueError("group_byクエリには'column'が必須です")
    return {
        "op": "group_by",
        "column": query["column"],
        "items": group_column_values(ws, query["column"], max_row),
    }


def _query_list_images(ws, query: dict, max_row: int) -> dict:
    items = []
    for idx, img in enumerate(ws._images):
        anchor = img.anchor
        if isinstance(anchor, AbsoluteAnchor):
            items.append(
                {
                    "image_index": idx,
                    "anchor_type": "absolute",
                    "left_cm": round(EMU_to_cm(anchor.pos.x), 2),
                    "top_cm": round(EMU_to_cm(anchor.pos.y), 2),
                    "width_cm": round(EMU_to_cm(anchor.ext.cx), 2),
                    "height_cm": round(EMU_to_cm(anchor.ext.cy), 2),
                }
            )
        else:
            items.append(
                {
                    "image_index": idx,
                    "anchor_type": type(anchor).__name__ if anchor is not None else None,
                    "left_cm": None,
                    "top_cm": None,
                    "width_cm": None,
                    "height_cm": None,
                }
            )
    return {"op": "list_images", "items": items}


def _chart_anchor_cell(anchor) -> str | None:
    """chart.anchorをセル参照文字列（例"D1"）へ変換する。

    メモリ上で`chart.anchor = "D1"`のように文字列で設定された場合はそのまま
    返す。保存済みファイルを読み込んだ場合は`OneCellAnchor`/`TwoCellAnchor`
    オブジェクトになるため、`_from`（0始まり行列）から変換する。
    """
    from openpyxl.utils import get_column_letter

    if isinstance(anchor, str):
        return anchor
    marker = getattr(anchor, "_from", None)
    if marker is None:
        return None
    try:
        return f"{get_column_letter(marker.col + 1)}{marker.row + 1}"
    except Exception:
        return None


def _chart_title_text(chart) -> str | None:
    """chart.titleからプレーンテキストを取り出す（タイトル未設定、または
    解釈できない構造の場合はNoneを返す）。"""
    title = chart.title
    if title is None:
        return None
    try:
        runs = title.tx.rich.p[0].r
        return "".join(r.t or "" for r in runs) or None
    except (AttributeError, IndexError, TypeError):
        return None


_CHART_TYPE_NAMES = {
    "LineChart": "line",
    "BarChart": "bar",
    "PieChart": "pie",
    "ScatterChart": "scatter",
}


def _query_list_charts(ws, query: dict, max_row: int) -> dict:
    items = []
    for idx, chart in enumerate(ws._charts):
        items.append(
            {
                "chart_index": idx,
                "type": _CHART_TYPE_NAMES.get(type(chart).__name__, type(chart).__name__),
                "title": _chart_title_text(chart),
                "anchor": _chart_anchor_cell(chart.anchor),
            }
        )
    return {"op": "list_charts", "items": items}


_QUERY_HANDLERS = {
    "group_by": _query_group_by,
    "list_images": _query_list_images,
    "list_charts": _query_list_charts,
}


def _run_queries(ws, queries: list[dict], max_row: int) -> list[dict]:
    results = []
    for query in queries:
        op_name = query.get("op")
        handler = _QUERY_HANDLERS.get(op_name)
        if handler is None:
            raise ValueError(f"未対応のqueryです: {op_name!r}（対応op: {sorted(_QUERY_HANDLERS)}）")
        results.append(handler(ws, query, max_row))
    return results


def _cell_json(cell) -> object:
    value = cell_to_json(cell.value)
    style = extract_style(cell)
    return {"value": value, "style": style} if style else {"value": value}


def _read_xlsx(
    path: Path,
    sheet_arg: str | None,
    offset: int,
    limit: int,
    data_only: bool,
    queries: list[dict] | None = None,
) -> dict:
    import openpyxl

    if queries and sheet_arg is None:
        raise ValueError("--query-json は --sheet 指定時のみ使えます")

    wb = openpyxl.load_workbook(str(path), data_only=data_only, read_only=False)
    try:
        names = wb.sheetnames
        if sheet_arg is None:
            sheets = [
                {"name": name, "max_row": wb[name].max_row or 0, "max_column": wb[name].max_column or 0}
                for name in names
            ]
            return {"path": str(path), "mode": "sheets", "sheets": sheets}

        resolved = resolve_sheet_name(names, sheet_arg)
        ws = wb[resolved]
        total_rows = ws.max_row or 0
        rows: list[list[object]] = []
        if total_rows:
            max_row = min(total_rows, offset + limit)
            if offset < max_row:
                for row in ws.iter_rows(min_row=offset + 1, max_row=max_row):
                    rows.append([_cell_json(cell) for cell in row])
        result = {
            "path": str(path),
            "mode": "rows",
            "sheet": resolved,
            "total_rows": total_rows,
            "total_columns": ws.max_column or 0,
            "start_row": offset + 1 if rows else None,
            "end_row": offset + len(rows) if rows else None,
            "rows": rows,
        }
        result["merged_cells"] = [str(r) for r in ws.merged_cells.ranges]
        result["tables"] = [
            {
                "name": t.name,
                "range": t.ref,
                "style": t.tableStyleInfo.name if t.tableStyleInfo else None,
            }
            for t in ws.tables.values()
        ]
        # 列幅を取得（返却範囲の列のみ）
        from openpyxl.utils import get_column_letter
        column_widths = {}
        if rows:
            min_col = 1
            max_col = len(rows[0]) if rows else 0
            for col_idx in range(min_col, min_col + max_col):
                letter = get_column_letter(col_idx)
                width = ws.column_dimensions[letter].width
                column_widths[letter] = width
        result["column_widths"] = column_widths

        # 行の高さを取得（返却範囲のみ）
        row_heights = {}
        if rows:
            for row_offset, row in enumerate(rows):
                row_num = offset + 1 + row_offset
                height = ws.row_dimensions[row_num].height
                if height is not None:
                    row_heights[str(row_num)] = height
        result["row_heights"] = row_heights

        # 列幅超過の警告を生成
        warnings = []
        for row_offset, row in enumerate(rows):
            row_num = offset + 1 + row_offset
            for col_idx, cell in enumerate(row):
                col_letter = get_column_letter(col_idx + 1)
                cell_width = column_widths.get(col_letter)
                # 各セルは _cell_json が {"value": ..., "style": {...}} の dict を返す
                cell_value = cell.get("value")
                wrap_text = cell.get("style", {}).get("wrap_text", False)
                # wrap_text が有効でない場合のみチェック
                if cell_width is not None and not wrap_text:
                    text_value = str(cell_value) if cell_value is not None else ""
                    if text_value:
                        display_width = _display_width(text_value)
                        # 列幅に2を加えたデフォルト幅（_apply_col_widths のロジックに合わせる）
                        target_width = min(max(cell_width + 2, 8), 60)
                        if display_width > target_width:
                            cell_ref = f"{resolved}!{col_letter}{row_num}"
                            msg = (
                                f"'{cell_ref}' の文字列(推定幅{display_width:.1f})が"
                                f"列幅({cell_width:.1f})を超えており、表示が切れる可能性があります"
                            )
                            warnings.append(msg)
                # 循環参照の警告を生成（数式が自身の座標を範囲に含む場合。
                # 例: N8セルに=SUM(B8:N8)（N8自身を含む）。--data-only指定時は
                # 数式文字列ではなくキャッシュ値が入るため自動的にスキップされる。
                if isinstance(cell_value, str) and cell_value.startswith("="):
                    if _formula_self_references(cell_value, col_idx + 1, row_num):
                        cell_ref = f"{resolved}!{col_letter}{row_num}"
                        warnings.append(
                            f"'{cell_ref}' の数式（{cell_value}）が自身のセルを範囲に含んでおり、"
                            "Excelで循環参照エラーになる可能性があります"
                        )
        if warnings:
            result["warnings"] = warnings
        if queries:
            result["query_results"] = _run_queries(ws, queries, total_rows)
        return result
    finally:
        wb.close()


def _read_xls(path: Path, sheet_arg: str | None, offset: int, limit: int) -> dict:
    import xlrd

    book = xlrd.open_workbook(str(path))
    names = book.sheet_names()
    if sheet_arg is None:
        sheets = []
        for name in names:
            sh = book.sheet_by_name(name)
            sheets.append({"name": name, "max_row": sh.nrows, "max_column": sh.ncols})
        return {"path": str(path), "mode": "sheets", "sheets": sheets}

    resolved = resolve_sheet_name(names, sheet_arg)
    sh = book.sheet_by_name(resolved)
    total_rows = sh.nrows
    end = min(total_rows, offset + limit)
    rows: list[list[object]] = []
    for r in range(offset, end):
        row_values = []
        for c in range(sh.ncols):
            cell = sh.cell(r, c)
            if cell.ctype == xlrd.XL_CELL_DATE:
                dt = xlrd.xldate.xldate_as_datetime(cell.value, book.datemode)
                row_values.append(dt.isoformat())
            elif cell.ctype == xlrd.XL_CELL_EMPTY:
                # xlrdは空セルをNoneではなく空文字列で返すため、SKILL.md記載通りnullに揃える
                row_values.append(None)
            else:
                row_values.append(cell.value)
        rows.append(row_values)
    return {
        "path": str(path),
        "mode": "rows",
        "sheet": resolved,
        "total_rows": total_rows,
        "total_columns": sh.ncols,
        "start_row": offset + 1 if rows else None,
        "end_row": offset + len(rows) if rows else None,
        "rows": rows,
    }


def main() -> int:
    setup_utf8_stdio()
    parser = argparse.ArgumentParser(allow_abbrev=False)
    parser.add_argument("file_path")
    parser.add_argument("--sheet", default=None, help="シート名または0始まりインデックス（省略時はシート一覧のみ返す）")
    parser.add_argument("--offset", type=int, default=0, help="読み飛ばす行数（0始まり、既定0）")
    parser.add_argument("--limit", type=int, default=200, help="読み込む最大行数（既定200）")
    parser.add_argument(
        "--data-only",
        action="store_true",
        help="xlsx/xlsmで数式ではなく最後にExcelが計算したキャッシュ値を返す（xlsのみ影響なし）",
    )
    parser.add_argument(
        "--query-json",
        default=None,
        help=(
            "構造化クエリの配列を1行JSON化した文字列（.xlsx/.xlsmのみ、--sheet必須）。"
            '例: \'[{"op": "group_by", "column": "A"}]\' '
            "→ 指定列を値ごとに連続する行範囲へグルーピングして{value,start_row,end_row,row_count}の"
            "配列で返す。insert_rows/merge_cells後の検証や、結合予定の列（月・区分等）の範囲確認を"
            "生のrowsを目で数えて手計算する代わりに使う。"
        ),
    )
    args = parser.parse_args()

    path = Path(args.file_path)
    if not path.is_file():
        print(f"ファイルが見つかりません: {args.file_path}", file=sys.stderr)
        return 1

    ext = path.suffix.lower()
    offset = max(args.offset, 0)
    limit = max(args.limit, 0)

    if ext == ".xls" and args.query_json:
        print("--query-json は .xlsx/.xlsm のみ対応です", file=sys.stderr)
        return 1

    queries: list[dict] | None = None
    if args.query_json:
        try:
            queries = json.loads(args.query_json)
        except json.JSONDecodeError as e:
            print(f"--query-jsonのJSON解析に失敗しました: {e}", file=sys.stderr)
            return 1
        if not isinstance(queries, list):
            print("--query-jsonはJSON配列で指定してください", file=sys.stderr)
            return 1

    try:
        if ext in (".xlsx", ".xlsm"):
            result = _read_xlsx(path, args.sheet, offset, limit, args.data_only, queries)
        elif ext == ".xls":
            result = _read_xls(path, args.sheet, offset, limit)
        else:
            print(f"未対応の拡張子です（.xlsx/.xlsm/.xls のみ対応）: {ext}", file=sys.stderr)
            return 1
    except ValueError as e:
        print(str(e), file=sys.stderr)
        return 1
    except ImportError as e:
        print(f"必要なライブラリが見つかりません: {e}", file=sys.stderr)
        return 1
    except Exception as e:  # noqa: BLE001 - 破損ファイル等、openpyxl/xlrdが送出する多様な例外を丸めて報告する
        print(f"読み込みに失敗しました: {e}\n{traceback.format_exc()}", file=sys.stderr)
        return 1

    summary = summarize_result(result, ["rows", "sheets"])
    summary.update(write_json_result(result, "excel_read", path))
    print(json.dumps(summary, ensure_ascii=False))
    return 0


if __name__ == "__main__":
    sys.exit(main())
